import os
import pandas
import csv
import glob
import openpyxl

glob.glob('*.xlsx') #拡張子が.xlsxのものを探索

def file_check(path):
    for pathname, dirnames, filenames in os.walk(path):
        for filename in filenames:
            print(os.path.join(pathname,filename))

def simple_check(path):
    i = 0
   
    for pathname,dirname, filenames in os.walk(path):
        for filename in filenames:
            # フィルタ処理 csvのみ処理
            if filename.endswith('.csv'):
                filename = os.path.join(*pathname,filename)
                # print(os.path.join(pathname,filename))
                print(filename)
                merge(filename,i)
                i += 1




def merge(csv_filename,i):
    
    # 拡張子なしのファイル名取得
    excel_filename_only = os.path.basename(excel_filename)
    csv_filename_only = os.path.splitext(csv_filename)[0].strip('/')

    # Creating a new excel file
    # wb = openpyxl.Workbook()
    csv_filename = Path.rstrip('/') + csv_filename


    print('csv-filename:'+ csv_filename)
    print('csv-filename-only:'+ csv_filename_only)

    # open a csv file
    with open(csv_filename,newline='') as csvf:
        # reading a csv file
        data = csv.reader(csvf)
        
        # Getting a new sheet
        wb.create_sheet(index = i, title=csv_filename_only)
        sheet = wb[csv_filename_only]


        r = 1
        for line in data:
            c = 1
            for v in line:
                sheet.cell(row = r,column = c).value = v
                c += 1
            r += 1


    # Saving a work book
    wb.save(filename = excel_filename_only)




# Locate file
excel_filename = './excel_test.xlsx'

# csvデータがあるところまでのPATH
Path = './sampledata/'

wb = openpyxl.Workbook()
simple_check(Path)
wb.close()
print('End')
glob.glob('*.xlsx') #拡張子が.xlsxのものを探索