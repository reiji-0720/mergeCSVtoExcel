import os
import csv
import openpyxl

# 作成したエクセルにCSVデータを反映させる
def merge(csv_filename,i,Path,wb):
    csv_filename_only = os.path.splitext(csv_filename)[0].strip('/') # CSVのファイル名だけ取り出す

    # Creating a new excel file
    # wb = openpyxl.Workbook()
    csv_filename = Path.rstrip('/') + csv_filename

    print('csv-filename:'+ csv_filename)
    print('csv-filename-only:'+ csv_filename_only)

    # open a csv file
    with open(csv_filename,newline='') as csvf:
        # reading a csv file
        data = csv.reader(csvf)

        # Getting a new sheet
        wb.create_sheet(index = i, title=csv_filename_only)
        sheet = wb[csv_filename_only]


        r = 1
        for line in data:
            c = 1 #colomn
            for v in line:
                sheet.cell(row = r,column = c).value = v
                c += 1
            r += 1 # row

        # columnのwidth設定
        # column_width = {'B': 3.5, 'C': 15, 'D': 20, 'E': 10, 'F': 10, 'G': 10}
        # for col, width in column_width.items():
        #     sheet.column_dimensions[col].width = width


        # セルのデータフォーマット指定（デバッグではうまくいくがエクセル側で問題がおきる）
        for row in sheet.iter_rows(min_col=1, max_col=6, min_row=2):
            row[0].number_format = 'yyyy-mm-dd'
            row[1].number_format = 'hh:mm:ss.0000'
            row[2].number_format = '#,##0'
            row[3].number_format = '0.000000'
            row[4].number_format = '0.000000'
            row[5].number_format = '0.000000'


        # sheet.cell(row=2,column=3).number_format = "#,##0"
        # sheet.cell(row=100,column=3).number_format = "#,##0"
        # sheet.cell(row=100,column=4).number_format = "0.000000"
        print(sheet.cell(row=2,column=1).number_format)
        print(sheet.cell(row=2,column=2).number_format)
        print(sheet.cell(row=2,column=3).number_format)
        print(sheet.cell(row=2,column=4).number_format)
        print(sheet.cell(row=2,column=5).number_format)
        print(sheet.cell(row=2,column=6).number_format)
